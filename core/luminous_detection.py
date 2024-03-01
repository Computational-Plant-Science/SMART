'''
Name: luminous_detection.py

Version: 1.0

Summary: Detect dark images by converting it to LAB color space to access the luminous channel which is independent of colors.
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2021-03-09

USAGE:

time python3 luminous_detection.py -p ~/plant-image-analysis/test/ -ft png 

'''

# import necessary packages
import argparse
import csv
from os.path import join
from typing import List

import cv2
import numpy as np
import os
import glob
from pathlib import Path

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from tabulate import tabulate
import openpyxl

# Convert it to LAB color space to access the luminous channel which is independent of colors.
from core.options import ImageInput


def isbright(options: ImageInput, threshold: float):
    # Load image file
    orig = cv2.imread(options.input_file)

    # Make backup image
    image = orig.copy()

    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    normalized = np.mean(L / np.max(L))

    # Normalize L channel by dividing all pixel values with maximum pixel value
    if normalized > threshold:
        text_bool = "bright"
        print(f"Image {options.input_stem} is light enough (normalized luminosity {normalized})")
    else:
        text_bool = "dark"
        print(f"Image {options.input_stem} is dark (normalized luminosity {normalized} under threshold {threshold})")

        # clahe = cv2.createCLAHE(clipLimit=8.0, tileGridSize=(3, 3))
        # cl = clahe.apply(L)
        # lightened = cv2.merge((cl, A, B))
        # converted = cv2.cvtColor(lightened, cv2.COLOR_LAB2BGR)

        # hsvImg = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        # hsvImg[..., 2] = hsvImg[..., 2] * 2
        # converted = cv2.cvtColor(hsvImg,cv2.COLOR_HSV2RGB)

        # hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        # h, s, v = cv2.split(hsv)
        # v += 255
        # final_hsv = cv2.merge((h, s, v))
        # converted = cv2.cvtColor(final_hsv, cv2.COLOR_HSV2BGR)

        # converted = increase_brightness(converted)

        # cv2.imwrite(join(options.output_directory, new_image_file), converted)

    return options.input_name, normalized, text_bool


def increase_brightness(img, value=150):
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)

    lim = 255 - value
    v[v > lim] = 255
    v[v <= lim] += value

    final_hsv = cv2.merge((h, s, v))
    img = cv2.cvtColor(final_hsv, cv2.COLOR_HSV2BGR)
    return img


def write_results_to_csv(results, output_directory):
    result_file = join(output_directory, 'luminous_detection.csv')
    headers = ['image_file_name', 'luminous_avg', 'dark_or_bright']

    with open(result_file, 'a+') as file:
        writer = csv.writer(file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        if os.stat(result_file).st_size == 0:
            writer.writerow(headers)

        for row in results:
            writer.writerow(row)


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required=True, help="path to image file")
    ap.add_argument("-ft", "--filetype", required=True, help="image filetype")
    ap.add_argument("-o", "--output_directory", required=True, help="directory to write output files to")

    args = vars(ap.parse_args())

    # Setting path to image files
    file_path = args["path"]
    file_type = args['filetype']
    output_dir = args['output_directory']

    # Extract file type and path
    filetype = '*.' + file_type
    image_file_path = file_path + filetype

    # Accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    options = [ImageInput(input_file=file, output_directory=output_dir) for file in imgList]

    # Get number of images in the data folder
    n_images = len(imgList)

    # get cpu number for parallel processing
    agents = psutil.cpu_count()
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))

    # Create a pool of processes. By default, one is created for each CPU in the machine.
    with closing(Pool(processes=agents)) as pool:
        results = pool.map(isbright, options)
        pool.terminate()

    # Output sum table in command window 
    print("Summary: {0} plant images were processed...\n".format(n_images))

    table = tabulate(results, headers=['image_file_name', 'luminous_avg', 'dark_or_bright'], tablefmt='orgtbl')

    print(table + "\n")

    write_results_to_csv(results, file_path)

import argparse
import cv2
import numpy as np
import os
import glob
from pathlib import Path

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from tabulate import tabulate
import openpyxl
from PIL import Image, ImageEnhance

import itertools

TEMPLATE_PATH = "/opt/spg-topdown-traits/marker_template/template.png"


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os

    # remove space at the beginning
    path = path.strip()
    # remove slash at the end
    path = path.rstrip("\\")

    # path exist?   # True  # False
    isExists = os.path.exists(path)

    # process
    if not isExists:
        # construct the path and folder
        # print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return
        # print path+' path exists!'
        return False


# get base name of a file from full path
def get_basename(image_file):
    abs_path = os.path.abspath(image_file)

    filename, file_extension = os.path.splitext(abs_path)

    base_name = os.path.splitext(os.path.basename(filename))[0]

    return base_name

    # image_file_name = Path(image_file).name


# Convert it to LAB color space to access the luminous channel which is independent of colors.
# def isbright(image_file):
#     # Set up threshold value for luminous channel, can be adjusted and generalized
#     thresh = 0.5
#
#     # Load image file
#     orig = cv2.imread(image_file)
#
#     # Make backup image
#     image = orig.copy()
#
#     # Get file name
#     # abs_path = os.path.abspath(image_file)
#
#     # filename, file_extension = os.path.splitext(abs_path)
#     # base_name = os.path.splitext(os.path.basename(filename))[0]
#
#     image_file_name = Path(image_file).name
#
#     # Convert color space to LAB format and extract L channel
#     L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
#
#     # Normalize L channel by dividing all pixel values with maximum pixel value
#     L = L / np.max(L)
#
#     text_bool = "bright" if np.mean(L) < thresh else "dark"
#
#     return image_file_name, np.mean(L), text_bool


# get weight value based on liner interpolation
def blend_weight_calculator(left_image_idx, right_image_idx, current_image_idx):
    window_width = right_image_idx - left_image_idx

    if window_width > 0:
        left_weight = abs(right_image_idx - current_image_idx) / window_width

        right_weight = abs(current_image_idx - left_image_idx) / window_width
    else:
        left_weight = 0.5
        right_weight = 0.5

    return left_weight, right_weight


# blend two images based on weights
def blend_image(left_image, right_image, left_weight, right_weight):
    left_img = cv2.imread(left_image)

    right_img = cv2.imread(right_image)

    blended = cv2.addWeighted(left_img, left_weight, right_img, right_weight, 0)

    return blended


# detect dark image and replac them with liner interpolated image
def check_discard_merge(options: List[ImageInput], replace: bool = False, threshold: float = 0.1):
    # create and assign index list for dark image
    idx_dark_imglist = [0] * len(options)

    result_list = []

    for idx, image in enumerate(options):
        img_name, mean_luminosity, luminosity_str = isbright(image.input_file,
                                                             threshold)  # luminosity detection, luminosity_str is either 'dark' or 'bright'
        result_list.append([img_name, mean_luminosity, luminosity_str])
        idx_dark_imglist[idx] = -1 if luminosity_str == 'dark' else (idx)

    table = tabulate(result_list, headers=['image_file_name', 'luminous_avg', 'dark_or_bright'], tablefmt='orgtbl')
    print(table + "\n")

    # save dark image detection result as excel file
    write_results_to_excel(result_list, options[0].output_directory)

    # print(idx_dark_imglist)

    # Finding consecutive occurrences of -1 in an array
    max_dark_list_length = max(len(list(v)) for g, v in itertools.groupby(idx_dark_imglist))

    # check max dark image sequence length, current method only deal with case with length equals 2
    # print(max_dark_list_length)

    # find dark image index
    idx_dark = [i for i, x in enumerate(idx_dark_imglist) if x == -1]
    idx_light = [i for i, x in enumerate(idx_dark_imglist) if x != -1]

    # print(idx_dark)

    # print(len(idx_dark_imglist))

    # process dark image
    if len(idx_dark) > 1:

        for idx, value in enumerate(idx_dark):

            # print("current value:{0}".format(value))

            # if dark image appears as the start of image list
            if value == 0:

                right_image_idx = ((value + 1) if idx_dark_imglist[value + 1] != -1 else (value + 2))

                left_image_idx = right_image_idx

            # if dark image appears as the end of image list
            elif value == len(idx_dark_imglist) - 1:

                left_image_idx = ((value - 1) if idx_dark_imglist[value - 1] != -1 else (value - 2))

                right_image_idx = left_image_idx

            else:

                left_image_idx = ((value - 1) if idx_dark_imglist[value - 1] != -1 else (value - 2))

                right_image_idx = ((value + 1) if idx_dark_imglist[value + 1] != -1 else (value + 2))

            # print("current image idx:{0}, left_idx:{1}, right_idx:{2}\n".format(value, left_image_idx, right_image_idx))

            (left_weight, right_weight) = blend_weight_calculator(left_image_idx, right_image_idx, value)

            # print("current image idx:{0}, left_idx:{1}, right_idx:{2}, left_weight:{3}, right_weight:{4} \n".format(value, left_image_idx, right_image_idx, left_weight, right_weight))

            blended = blend_image(options[left_image_idx].input_file, options[right_image_idx].input_file, left_weight, right_weight)

            print("Blending image:{0}, left:{1}, right:{2}, left_weight:{3:.2f}, right_weight:{4:.2f}".format(options[value].input_stem,
                                                                                                              options[left_image_idx].input_stem,
                                                                                                              options[right_image_idx].input_stem,
                                                                                                              left_weight, right_weight))

            # save result by overwriting original files
            # cv2.imwrite(options[value].input_file, blended)

            # save result into result folder for debugging
            cv2.imwrite(join(options[0].output_directory,
                             options[value].input_file) if replace else f"{join(options[0].output_directory, options[value].input_stem)}.blended.png",
                        blended)

    for idx, value in enumerate(idx_light):
        image = cv2.imread(options[value].input_file)
        cv2.imwrite(join(options[0].output_directory,
                         options[value].input_file) if replace else f"{join(options[0].output_directory, options[value].input_stem)}.png", image)


def check_discard_merge2(options: List[ImageInput], threshold: float = 0.1):
    left = None
    right = None
    i = 0
    replaced = 0
    any_dark = False

    # if every image has timestamp data, sort images by timestamp
    if all(option.timestamp is not None for option in options):
        options = sorted(options, key=lambda o: o.timestamp)

    for option in options:
        img_name, mean_luminosity, luminosity_str = isbright(option, threshold)  # luminosity detection, luminosity_str is either 'dark' or 'bright'
        write_results_to_csv([(img_name, mean_luminosity, luminosity_str)], option.output_directory)
        if luminosity_str == 'dark':
            print(f"{option.input_stem} is too dark, skipping")
            any_dark = True
            continue
        else:
            path = join(options[0].output_directory, Path(option.input_file).name)
            print(f"Writing to {path}")
            cv2.imwrite(path, cv2.imread(option.input_file))
        # if luminosity_str == 'dark':
        #     if left is None:
        #         left = i
        #     right = i
        #     any_dark = True
        # elif left is not None and left != right:
        #     ii = 0
        #     left_file = sorted_options[left].input_file
        #     right_file = sorted_options[right].input_file
        #     prev_file = sorted_options[left - 1].input_file
        #     next_file = sorted_options[right + 1].input_file
        #     prev = cv2.imread(prev_file)
        #     next = cv2.imread(next_file)
        #     width = right - left + 1
        #     offset = (1 / width)
        #     print(f"Replacing {left_file} to {right_file} with merger of {prev_file} and {next_file}")
        #     for opt in reversed(sorted_options[left:right + 1]):
        #         prev_weight = ((ii / (right - left)) * ((width - 1) / width)) + offset
        #         next_weight = ((1 - prev_weight) * ((width - 1) / width)) + offset
        #         print(f"Merging {prev_file} (weight {round(prev_weight, 2)}) with {next_file} (weight: {round(next_weight, 2)})")
        #         blended = cv2.addWeighted(prev, prev_weight, next, next_weight, 0)
        #         # cv2.imwrite(opt.input_file, blended)
        #         cv2.imwrite(join(options[0].output_directory, opt.input_file) if replace else f"{join(options[0].output_directory, opt.input_stem)}.blended.png", blended)
        #         ii += 1
        #     left = None
        #     right = None
        #     replaced += width
        # else:
        #     cv2.imwrite(join(options[0].output_directory, option.input_file) if replace else f"{join(options[0].output_directory, option.input_stem)}.png", cv2.imread(option.input_file))
        # i += 1
    # print(f"Replaced {replaced} dark images with weighted blends of adjacent images")
    return any_dark


# Detect circles in the image
def circle_detect(image_path, template_path):
    print(f"Checking for circle to crop in {image_path}")
    template = cv2.imread(template_path, 0)

    # load the image, clone it for output, and then convert it to grayscale
    img_ori = cv2.imread(image_path)

    img_rgb = img_ori.copy()

    # Convert it to grayscale
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

    # Store width and height of template in w and h
    w, h = template.shape[::-1]

    # Perform match operations.
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)

    # Specify a threshold
    threshold = 0.8

    # Store the coordinates of matched area in a numpy array
    loc = np.where(res >= threshold)

    if len(loc):

        (y, x) = np.unravel_index(res.argmax(), res.shape)

        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)

        # print(y,x)

        # print(min_val, max_val, min_loc, max_loc)

        # Draw a rectangle around the matched region.
        for pt in zip(*loc[::-1]):
            circle_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0, 255, 255), 2)

            # save segmentation result
        # result_file = (save_path + base_name + '_circle.' + args['filetype'])
        # print(result_file)
        # cv2.imwrite(result_file, circle_overlay)

        crop_img = img_rgb[y + 150:y + 850, x - 650:x]

        # save segmentation result
        # result_file = (save_path + base_name + '_cropped.' + args['filetype'])
        # print(result_file)
        # cv2.imwrite(result_file, crop_img)

    return crop_img


def image_enhance(image_file):
    im = Image.fromarray(cv2.cvtColor(image_file, cv2.COLOR_BGR2RGB))

    im_sharpness = ImageEnhance.Sharpness(im).enhance(3.5)

    im_contrast = ImageEnhance.Contrast(im_sharpness).enhance(1.5)

    im_out = ImageEnhance.Brightness(im_contrast).enhance(1.2)

    return im_out


def write_results_to_csv(results, output_directory):
    result_file = join(output_directory, 'luminous_detection.csv')
    headers = ['image_file_name', 'luminous_avg', 'dark_or_bright']

    with open(result_file, 'a+') as file:
        writer = csv.writer(file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        if os.stat(result_file).st_size == 0:
            writer.writerow(headers)

        for row in results:
            writer.writerow(row)


def write_results_to_excel(results, output_directory):
    result_file = join(output_directory, 'luminous_detection.xlsx')

    # print(result_file)

    if os.path.isfile(result_file):
        # update values
        # Open an xlsx for reading
        wb = openpyxl.load_workbook(result_file)

        # Get the current Active Sheet
        sheet = wb.active

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.cell(row=1, column=1).value = 'image_file_name'
        sheet.cell(row=1, column=2).value = 'luminous_avg'
        sheet.cell(row=1, column=3).value = 'dark_or_bright'

    for row in results:
        sheet.append(row)

    # save the csv file
    wb.save(result_file)


if __name__ == '__main__':

    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required=True, help="path to image file")
    ap.add_argument("-ft", "--filetype", required=True, help="image filetype")

    args = vars(ap.parse_args())

    # Setting path to image files
    file_path = args["path"]
    ext = args['filetype']

    # Extract file type and path
    filetype = '*.' + ext
    image_file_path = file_path + filetype

    # Accquire image file list
    imgList = sorted(glob.glob(image_file_path))

    mkpath = os.path.dirname(file_path) + '/merged'
    mkdir(mkpath)
    save_path = mkpath + '/'

    # Read the template
    template = cv2.imread(TEMPLATE_PATH, 0)

    # print((imgList))
    # global save_path

    # Get number of images in the data folder
    n_images = len(imgList)

    # replace dark image using blended image
    check_discard_merge(imgList)

    # enhance image Contrast,Brightness,Sharpness
    for image in imgList:
        im_out_name = get_basename(image)

        # construct the result file path
        result_img_path = save_path + im_out_name + '.' + ext

        crop_img = circle_detect(image)

        im_out = image_enhance(crop_img)

        im_out.save(result_img_path)
